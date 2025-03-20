import pandas as pd
import os
import openpyxl


REPORTS_PATH = "./files/reports/"

RATINGS = "PlayerRatings"
CURRENT_PIVOTS = "CurrentPivots"
PLAYER_BATTING = "PlayerBatting"
PLAYER_PITCHING = "PlayerPitching"
LEAGUE_BATTING = "LeagueBatting"
LEAGUE_PITCHING = "LeaguePitching"
PLAYER_FIELDING = "PlayerFielding"
LEAGUE_FIELDING = "LeagueFielding"
WOBA = "wOBACalcs"
FIELDING_PLAYABLES = "FieldingPlayables"
PREDICTIONS = "Predictions"
POTENTIALS = "Potentials"

TABS = [
    RATINGS,
    CURRENT_PIVOTS,
    PLAYER_BATTING,
    PLAYER_PITCHING,
    LEAGUE_BATTING,
    LEAGUE_PITCHING,
    PLAYER_FIELDING,
    LEAGUE_FIELDING,
    WOBA,
    FIELDING_PLAYABLES,
    PREDICTIONS,
    POTENTIALS,
]


class ExcelReportWriter:

    def __init__(self, league: str):
        self.league = league
        self.filename = f"{REPORTS_PATH}/{league}.xlsx"

    def clear_workbook(self, sheet_name: str):

        wb = openpyxl.load_workbook(filename=self.filename)
        for ws in wb.worksheets:
            if ws.title == sheet_name:
                for row in ws.iter_rows(
                    min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column
                ):
                    for cell in row:
                        cell.value = None

        wb.save(self.filename)

    def create_summary(
        self,
        ratings: pd.DataFrame,
        batting_stats: pd.DataFrame,
        pitching_stats: pd.DataFrame,
        fielding_stats: pd.DataFrame,
    ) -> pd.DataFrame:

        df_summary = ratings.copy(deep=True)
        df_summary = df_summary[["ID", "LPOS", "First Name", "Last Name", "ORG", "Age"]]

        df_batting_report = batting_stats[["ID", "wRAA", "BSR", "OFF", "OFFAdj"]]
        df_summary = pd.merge(df_summary, df_batting_report, on="ID", how="left")

        fielding_group = fielding_stats.groupby("ID")["runsPAdj"].sum().reset_index()
        df_summary = pd.merge(
            df_summary, fielding_group, on="ID", how="left", suffixes=("", "_Sum")
        )

        df_pitching_report = pitching_stats[
            ["ID", "IPClean", "WAA", "WAAAdj", "WAA200"]
        ]
        df_summary = pd.merge(df_summary, df_pitching_report, on="ID", how="left")

        df_summary = df_summary[
            (df_summary["wRAA"].notna()) | (df_summary["IPClean"].notna())
        ]
        df_summary.fillna(-500.0, inplace=True)

        df_summary = df_summary[
            [
                "ID",
                "LPOS",
                "First Name",
                "Last Name",
                "ORG",
                "Age",
                "wRAA",
                "BSR",
                "OFF",
                "OFFAdj",
                "runsPAdj",
                "IPClean",
                "WAA",
                "WAAAdj",
                "WAA200",
            ]
        ]

        return df_summary

    def load_player_stats(self, season: int):

        file_exists = os.path.exists(self.filename)

        if file_exists:
            for tab in TABS:
                self.clear_workbook(tab)

        ratings = pd.read_csv(
            f"./files/{self.league}/{season}/output/{self.league}-{season}-player-data.csv"
        )
        batting_stats = pd.read_csv(
            f"./files/{self.league}/{season}/output/{self.league}-{season}-hitting.csv"
        )
        pitching_stats = pd.read_csv(
            f"./files/{self.league}/{season}/output/{self.league}-{season}-pitching.csv"
        )
        fielding_stats = pd.read_csv(
            f"./files/{self.league}/{season}/output/{self.league}-{season}-fielding.csv"
        )
        league_batting_stats = pd.read_csv(
            f"./files/{self.league}/{season}/output/{self.league}-{season}-team-batting.csv"
        )
        league_pitching_stats = pd.read_csv(
            f"./files/{self.league}/{season}/output/{self.league}-{season}-team-pitching.csv"
        )
        league_fielding_stats = pd.read_csv(
            f"./files/{self.league}/{season}/output/{self.league}-{season}-team-fielding.csv"
        )
        woba = pd.read_csv(
            f"./files/{self.league}/{season}/output/{self.league}-{season}-woba-calcs.csv"
        )
        fielding_playables = pd.read_csv(
            f"./files/{self.league}/{season}/output/{self.league}-{season}-fielding-attributes.csv"
        )

        # FIXME
        batting_stats = batting_stats.drop(columns=["BsR"])
        batting_stats.rename(columns={"BSR2": "BSR"}, inplace=True)

        observed_summary = self.create_summary(
            ratings, batting_stats, pitching_stats, fielding_stats
        )

        args = {
            "path": self.filename,
            "engine": "openpyxl",
            "mode": "a" if file_exists else "w",
        }

        args = args | ({"if_sheet_exists": "overlay"} if file_exists else {})

        with pd.ExcelWriter(**args) as writer:

            observed_summary.to_excel(writer, sheet_name=CURRENT_PIVOTS, index=False)
            ratings.to_excel(writer, sheet_name=RATINGS, index=False)
            batting_stats.to_excel(writer, sheet_name=PLAYER_BATTING, index=False)
            pitching_stats.to_excel(writer, sheet_name=PLAYER_PITCHING, index=False)
            fielding_stats.to_excel(writer, sheet_name=PLAYER_FIELDING, index=False)
            league_batting_stats.to_excel(
                writer, sheet_name=LEAGUE_BATTING, index=False
            )
            league_pitching_stats.to_excel(
                writer, sheet_name=LEAGUE_PITCHING, index=False
            )
            league_fielding_stats.to_excel(
                writer, sheet_name=LEAGUE_FIELDING, index=False
            )
            woba.to_excel(writer, sheet_name=WOBA, index=False)
            fielding_playables.to_excel(
                writer, sheet_name=FIELDING_PLAYABLES, index=False
            )
            # Styler.to_excel(writer, sheet_name=RATINGS, float_format="%.3f")

        wb = openpyxl.load_workbook(filename=self.filename)
        if CURRENT_PIVOTS not in wb[CURRENT_PIVOTS].tables:
            tab = openpyxl.worksheet.table.Table(
                displayName=CURRENT_PIVOTS,
                ref=f"A1:{openpyxl.utils.get_column_letter(ratings.shape[1])}{len(ratings)+1}",
            )
            wb[CURRENT_PIVOTS].add_table(tab)
            wb.save(self.filename)

        if RATINGS not in wb[RATINGS].tables:
            tab = openpyxl.worksheet.table.Table(
                displayName=RATINGS,
                ref=f"A1:{openpyxl.utils.get_column_letter(ratings.shape[1])}{len(ratings)+1}",
            )
            wb[RATINGS].add_table(tab)
            wb.save(self.filename)

        if PLAYER_BATTING not in wb[PLAYER_BATTING].tables:
            tab = openpyxl.worksheet.table.Table(
                displayName=PLAYER_BATTING,
                ref=f"A1:{openpyxl.utils.get_column_letter(batting_stats.shape[1])}{len(batting_stats)+1}",
            )
            wb[PLAYER_BATTING].add_table(tab)
            wb.save(self.filename)

        if PLAYER_PITCHING not in wb[PLAYER_PITCHING].tables:
            tab = openpyxl.worksheet.table.Table(
                displayName=PLAYER_PITCHING,
                ref=f"A1:{openpyxl.utils.get_column_letter(pitching_stats.shape[1])}{len(pitching_stats)+1}",
            )
            wb[PLAYER_PITCHING].add_table(tab)
            wb.save(self.filename)

        if PLAYER_FIELDING not in wb[PLAYER_FIELDING].tables:
            tab = openpyxl.worksheet.table.Table(
                displayName=PLAYER_FIELDING,
                ref=f"A1:{openpyxl.utils.get_column_letter(fielding_stats.shape[1])}{len(fielding_stats)+1}",
            )
            wb[PLAYER_FIELDING].add_table(tab)
            wb.save(self.filename)

        if LEAGUE_FIELDING not in wb[LEAGUE_FIELDING].tables:
            tab = openpyxl.worksheet.table.Table(
                displayName=LEAGUE_FIELDING,
                ref=f"A1:{openpyxl.utils.get_column_letter(league_fielding_stats.shape[1])}{len(league_fielding_stats)+1}",
            )
            wb[LEAGUE_FIELDING].add_table(tab)
            wb.save(self.filename)

        if LEAGUE_BATTING not in wb[LEAGUE_BATTING].tables:
            tab = openpyxl.worksheet.table.Table(
                displayName=LEAGUE_BATTING,
                ref=f"A1:{openpyxl.utils.get_column_letter(league_batting_stats.shape[1])}{len(league_batting_stats)+1}",
            )
            wb[LEAGUE_BATTING].add_table(tab)
            wb.save(self.filename)

        if LEAGUE_PITCHING not in wb[LEAGUE_PITCHING].tables:
            tab = openpyxl.worksheet.table.Table(
                displayName=LEAGUE_PITCHING,
                ref=f"A1:{openpyxl.utils.get_column_letter(league_pitching_stats.shape[1])}{len(league_pitching_stats)+1}",
            )
            wb[LEAGUE_PITCHING].add_table(tab)
            wb.save(self.filename)

        if WOBA not in wb[WOBA].tables:
            tab = openpyxl.worksheet.table.Table(
                displayName=WOBA,
                ref=f"A1:{openpyxl.utils.get_column_letter(woba.shape[1])}{len(woba)+1}",
            )
            wb[WOBA].add_table(tab)
            wb.save(self.filename)

        if FIELDING_PLAYABLES not in wb[FIELDING_PLAYABLES].tables:
            tab = openpyxl.worksheet.table.Table(
                displayName=FIELDING_PLAYABLES,
                ref=f"A1:{openpyxl.utils.get_column_letter(fielding_playables.shape[1])}{len(fielding_playables)+1}",
            )
            wb[FIELDING_PLAYABLES].add_table(tab)
            wb.save(self.filename)

        wb.close()

    def write_predictions(self, df_predictions: pd.DataFrame):

        args = {
            "path": self.filename,
            "engine": "openpyxl",
            "mode": "a",
            "if_sheet_exists": "overlay",
        }

        with pd.ExcelWriter(**args) as writer:
            df_predictions.to_excel(writer, sheet_name=PREDICTIONS, index=False)

        wb = openpyxl.load_workbook(filename=self.filename)
        if PREDICTIONS not in wb[PREDICTIONS].tables:
            tab = openpyxl.worksheet.table.Table(
                displayName=PREDICTIONS,
                ref=f"A1:{openpyxl.utils.get_column_letter(df_predictions.shape[1])}{len(df_predictions)+1}",
            )
            wb[PREDICTIONS].add_table(tab)
            wb.save(self.filename)

        wb.close()

    def write_potential_predictions(self, df_predictions: pd.DataFrame):

        args = {
            "path": self.filename,
            "engine": "openpyxl",
            "mode": "a",
            "if_sheet_exists": "overlay",
        }

        with pd.ExcelWriter(**args) as writer:
            df_predictions.to_excel(writer, sheet_name=POTENTIALS, index=False)

        wb = openpyxl.load_workbook(filename=self.filename)
        if POTENTIALS not in wb[POTENTIALS].tables:
            tab = openpyxl.worksheet.table.Table(
                displayName=POTENTIALS,
                ref=f"A1:{openpyxl.utils.get_column_letter(df_predictions.shape[1])}{len(df_predictions)+1}",
            )
            wb[POTENTIALS].add_table(tab)
            wb.save(self.filename)

        wb.close()
